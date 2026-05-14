import pandas as pd
import numpy as np
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
import os

# ============================================================
# CONFIGURATION
# ============================================================

LEAD_TIME_DAYS = 4
OUTPUT_FILE = 'output/Inventory_Reorder_Report.xlsx'

# Colors
COLOR_HEADER     = '1F4E79'   # Dark blue for column headers
COLOR_URGENT     = 'C00000'   # Dark red
COLOR_LOW        = 'C55A11'   # Dark amber
COLOR_MONITOR    = '375623'   # Dark green
COLOR_ROW_ALT    = 'EBF3FB'   # Light blue alternate row
COLOR_WHITE      = 'FFFFFF'


# ============================================================
# HELPERS
# ============================================================

def reorder_date(lead_time_days=LEAD_TIME_DAYS):
    return (datetime.today() + timedelta(days=lead_time_days)).strftime('%d/%m/%Y')


def thin_border():
    side = Side(style='thin', color='BFBFBF')
    return Border(left=side, right=side, top=side, bottom=side)


def header_border():
    side = Side(style='medium', color='FFFFFF')
    return Border(left=side, right=side, top=side, bottom=side)

def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

DB_PATH = os.path.join(get_base_path(), 'inventory_optimizer.db')

# ============================================================
# GENERATE REPORT
# ============================================================

def generate_report(results_df, all_results_df, output_path=OUTPUT_FILE):
    """
    Generates a clean Excel report.
    Starts directly with column headers and data.
    No summary block inside the file.

    Parameters:
        results_df:     items requiring reorder
        all_results_df: all items analyzed
        output_path:    save location
    """

    base = get_base_path()
    output_dir = os.path.join(base, 'output')
    os.makedirs(output_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reorder Recommendations'

    order_by = reorder_date()

    # --------------------------------------------------------
    # COLUMN DEFINITIONS
    # --------------------------------------------------------

    columns = [
        ('Item Code',                15),
        ('Item Description',         40),
        ('Unit of Measure',          16),
        ('Forecasted Demand\n(Next Month)', 18),
        ('Current Stock',            14),
        ('Recommended Order',        18),
        ('Urgency',                  12),
        ('Reorder By',               14),
    ]

    # --------------------------------------------------------
    # HEADER ROW
    # --------------------------------------------------------

    ws.row_dimensions[1].height = 32

    for col_idx, (header, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(
            name='Calibri', size=10,
            bold=True, color=COLOR_WHITE
        )
        cell.fill = PatternFill(
            fill_type='solid', fgColor=COLOR_HEADER
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        cell.border = header_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --------------------------------------------------------
    # SORT DATA
    # --------------------------------------------------------

    # Urgent items (zero stock) first, then by lowest stock
    results_sorted = results_df.copy()
    results_sorted['_sort'] = results_sorted['Current_Stock'].apply(
        lambda x: 0 if x <= 0 else 1
    )
    results_sorted = results_sorted.sort_values(
        ['_sort', 'Current_Stock']
    ).drop(columns=['_sort']).reset_index(drop=True)

    # DATA ROWS

    for row_idx, row in results_sorted.iterrows():

        excel_row = row_idx + 2
        ws.row_dimensions[excel_row].height = 15

        current  = row['Current_Stock']
        forecast = row['Forecasted_Demand']

        # Urgency logic
        if current <= 0:
            urgency       = 'URGENT'
            urgency_color = COLOR_URGENT
        elif current < forecast * 0.3:
            urgency       = 'LOW'
            urgency_color = COLOR_LOW
        else:
            urgency       = 'MONITOR'
            urgency_color = COLOR_MONITOR

        # Alternating row color
        row_bg = COLOR_ROW_ALT if row_idx % 2 == 0 else COLOR_WHITE

        row_data = [
            row['Item Code'],
            row['Item Description'],
            row['Unit of Measure'],
            round(forecast, 0),
            round(current, 0),
            round(row['Recommended_Order'], 0),
            urgency,
            order_by,
        ]

        for col_idx, value in enumerate(row_data, start=1):

            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = Font(name='Calibri', size=9)
            cell.border = thin_border()

            # Item Description left aligned, rest centered
            cell.alignment = Alignment(
                horizontal='left' if col_idx == 2 else 'center',
                vertical='center',
                indent=1 if col_idx == 2 else 0
            )

            # Urgency column: colored text, no background fill
            if col_idx == 7:
                cell.font = Font(
                    name='Calibri', size=9,
                    bold=True, color=urgency_color
                )
                cell.fill = PatternFill(
                    fill_type='solid', fgColor=row_bg
                )
            else:
                cell.fill = PatternFill(
                    fill_type='solid', fgColor=row_bg
                )
    # FREEZE TOP ROW

    ws.freeze_panes = 'A2'
    # AUTO FILTER ON HEADER ROW

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"
    # SAVE
    timestamp = datetime.today().strftime('%Y-%m-%d_%H%M')
    output_path = os.path.join(output_dir, f'Inventory_Reorder_Report_{timestamp}.xlsx')
    wb.save(output_path)

    print(f"\nReport saved: {output_path}")
    print(f"Items in report: {len(results_df)}")

    return output_path


if __name__ == '__main__':
    from preprocessing import run_preprocessing
    from safety_stock import build_safety_stock_table, get_current_stock
    from forecast import forecast_all_items
    from milp_model import optimize_all_items

    full_df, monthly_df   = run_preprocessing()
    reorder_df            = build_safety_stock_table(full_df)
    current_stock_df      = get_current_stock(full_df)
    forecast_df           = forecast_all_items(full_df)
    results_df, all_results_df = optimize_all_items(
        forecast_df, reorder_df, current_stock_df
    )

    generate_report(results_df, all_results_df)